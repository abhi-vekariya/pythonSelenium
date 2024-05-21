import openpyxl

file = '/home/abhivekariya/pythonSelenium/Data_driven_testing/abcd1.xlsx'

workbook = openpyxl.load_workbook(file)
sheet = workbook.active #or we can write workbook["sheet1"]

# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value='abhi'
#
# workbook.save(file)

sheet.cell(1,1).value=123
sheet.cell(1,2).value='Mohan'

sheet.cell(2,1).value=456
sheet.cell(2,2).value='Rohan'

sheet.cell(3,1).value=789
sheet.cell(3,2).value='Kishan'

workbook.save(file)



