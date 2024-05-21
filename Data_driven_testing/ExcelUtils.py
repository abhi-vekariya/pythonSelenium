import openpyxl
from openpyxl.styles import PatternFill

# Function to get the number of rows in a specified sheet of an Excel file
def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

# Function to get the number of columns in a specified sheet of an Excel file
def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

# Function to read data from a specific cell in a specified sheet of an Excel file
def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno).value

# Function to write data to a specific cell in a specified sheet of an Excel file
def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,columnno).value = data
    workbook.save(file)

# Function to fill a specific cell with green color in a specified sheet of an Excel file
def fillGreenColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum,columnno).fill = greenFill
    workbook.save(file)

# Function to fill a specific cell with red color in a specified sheet of an Excel file
def fillRedColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownum,columnno).fill = redFill
    workbook.save(file)

